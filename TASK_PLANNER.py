
#######################################################################################################################
from PyQt5.QtWidgets import QMessageBox, QLineEdit
from PyQt5.uic import loadUi
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5 import QtGui
import openpyxl
######################################################################################################################
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import keyboard
import time
import os
import shutil
########################################################################################################################
class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        loadUi("taskplanner.ui", self)
        self.comboBox.setItemText(0, "IMS")
        self.comboBox.setItemText(1, "Virtualization")
        self.comboBox_2.setItemText(0, "TT")
        self.comboBox_2.setItemText(1, "NC")
        self.comboBox_2.setItemText(2, "TTNET")
        self.setWindowTitle("Netas Task Planner")
        self.pushButton.clicked.connect(self.namesurname)
        self.comboBox.activated[str].connect(self.on_combobox_activated)
        self.comboBox_2.activated[str].connect(self.company)
        self.pushButton_4.clicked.connect(self.time)
        self.calendarWidget.clicked.connect(self.on_calendar_clicked)
        self.pushButton_2.clicked.connect(self.on_button_clicked)
        self.pushButton_3.clicked.connect(self.clear)
        self.pushButton_7.clicked.connect(self.login)
        self.lineEdit_3.setEchoMode(QLineEdit.Password)
        self.pushButton_5.clicked.connect(self.upload)
        self.setFixedSize(889, 722)
        self.comboBox = None
        self.isim = None
        self.tarih = None
        self.zaman = None
        self.combobox2 = None
        self.email = None
        self.password = None
        ################################################################################################################

        ################################################################################################################
    def login(self):
        QMessageBox.information(self, "Sharepoint Information", "Do not close the browser until the file has been downloaded.")
        driver = webdriver.Chrome()
        driver.get("https://nortel365.sharepoint.com/sites/CORENETWORKORTAKALANI/Shared Documents/General")
        email = self.lineEdit_2.text()
        şifre = self.lineEdit_3.text()
        ##############################
        time.sleep(2)
        email_field = driver.find_element(By.CSS_SELECTOR, "input[name='loginfmt']")
        email_field.send_keys(email)
        email_field.send_keys(Keys.RETURN)
        ##############################
        time.sleep(2)
        password_field = driver.find_element(By.CSS_SELECTOR, "input[name='passwd']")
        password_field.send_keys(şifre)
        password_field.send_keys(Keys.RETURN)
        #############################
        time.sleep(2)
        text_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, "//*[text()='Approve a request on my Microsoft Authenticator app']"))
        )
        text_element.click()
        #############################
        if EC.presence_of_element_located((By.ID, "idBtn_Back")):
             print("5")
             back_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "idBtn_Back"))
             )
             back_button.click()
        ##############################
        time.sleep(2)
        text_element2 = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[text()='Planlı Çalışma-Freeze Takvimi.xlsx']"))
        )
        actions = ActionChains(driver)
        actions.context_click(text_element2).perform()
        ###############################
        time.sleep(2)
        text_element3 = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[text()='Download']"))
        )
        text_element3.click()
        ##############################
        time.sleep(5)
        kullanici_klasoru = os.path.expanduser('~')
        indirilenler_klasoru = os.path.join(kullanici_klasoru, 'Downloads')
        dosya_adı = 'Planlı Çalışma-Freeze Takvimi.xlsx'
        dosya_yolu = os.path.join(indirilenler_klasoru, dosya_adı)
        hedef_yolu = "C:/TASK PLANNER/"
        hedef_dosya_yolu = os.path.join(hedef_yolu, dosya_adı)
        if os.path.exists(hedef_dosya_yolu):
            os.remove(hedef_dosya_yolu)
        shutil.copy2(dosya_yolu, hedef_yolu)
        os.remove(dosya_yolu)
        ################################################
    def namesurname(self):
        self.isim = self.lineEdit.text()
        QMessageBox.information(self, "Select Your Team And Company", "Select team information and company")

    def on_combobox_activated(self, text):
        self.comboBox = text
       

    def time(self):
        self.zaman = self.timeEdit.text()

    def company(self,text):
        self.combobox2 = text



    def on_calendar_clicked(self, date):
        self.tarih = date.toString("dd.MM.yyyy")

    def on_button_clicked(self):
        text = self.textEdit.toPlainText()
        if self.checkBox.isChecked() and self.checkBox_2.isChecked() and self.checkBox_3.isChecked():
            text += ' - ZTE Ticket'
            text += ' - Mop'
            text += ' - Customer Approval'
        ########################################################
        workbook = openpyxl.load_workbook('Planlı Çalışma-Freeze Takvimi.xlsx')
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        tarih_hucre = None
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for column in worksheet.iter_cols(min_col=1, max_col=len(columns), min_row=1):
                for cell in column:
                    if cell.value == self.tarih:
                        tarih_hucre = cell
                        break
                if tarih_hucre:
                    break
            if tarih_hucre:
                break
        if not tarih_hucre:
            QMessageBox.warning(self, "Operation Check", "No Date Available.")
        ###########################################################################################################
        else:  #
            if self.checkBox.isChecked() and self.checkBox_2.isChecked() and self.checkBox_3.isChecked():  #
                if self.combobox2:  #
                    ###############################################################################################
                    if self.comboBox != "Virtualization":
                        satir = tarih_hucre.row + 1
                        sutun = tarih_hucre.column
                        value = worksheet.cell(row=satir, column=sutun).value
                        if value is None or len(value) == 1:
                            worksheet.cell(row=satir, column=sutun, value=(text + ' - ' + self.zaman + ' - ' + ' (' + self.isim + ')' + ' - ' + self.comboBox + ' - ' + 'Company: ' + self.combobox2 + ' - '))
                            workbook.save('Planlı Çalışma-Freeze Takvimi.xlsx')
                            QMessageBox.information(self, "Mission status", "The task has been created.")
                        else:
                            secim = QMessageBox.question(self, "Task Exists", "A task already exists on this date. Do you want to add a new task to this date?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                            if secim == QMessageBox.Yes:
                                worksheet.cell(row=satir, column=sutun, value=(value + '\n' + '\n' + text + ' - ' + self.zaman + ' - ' + ' (' + self.isim + ')' + ' - ' + self.comboBox + ' - ' + 'Company: ' + self.combobox2 + ' - '))
                                workbook.save('Planlı Çalışma-Freeze Takvimi.xlsx')
                                QMessageBox.information(self, "Mission status", "The task has been added.")
                            else:
                                QMessageBox.information(self, "Operation Cancelled", "The operation has been cancelled.")

                    elif self.comboBox == "Virtualization":
                        satir = tarih_hucre.row + 2
                        sutun = tarih_hucre.column
                        value = worksheet.cell(row=satir, column=sutun).value
                        if value is None or len(value) == 1:
                            worksheet.cell(row=satir, column=sutun, value=(text + ' - ' + self.zaman + ' - ' + ' (' + self.isim + ')' + ' - ' + 'Company: ' + self.combobox2 + ' - '))
                            workbook.save('Planlı Çalışma-Freeze Takvimi.xlsx')
                            QMessageBox.information(self, "Mission status", "The task has been created.")
                        else:
                            secim = QMessageBox.question(self, "Task Exists","A task already exists on this date. Do you want to add a new task to this date?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                            if secim == QMessageBox.Yes:
                                worksheet.cell(row=satir, column=sutun, value=( value + '\n' + '\n' + text + ' - ' + self.zaman + ' - ' + ' (' + self.isim + ')' + ' - ' + 'Company: ' + self.combobox2 + ' - '))
                                workbook.save('Planlı Çalışma-Freeze Takvimi.xlsx')
                                QMessageBox.information(self, "Mission status", "The task has been added.")
                            else:
                                QMessageBox.information(self, "Operation Cancelled", "The operation has been cancelled.")
             ############################################################################################################################
            else:
                #################################################################################################
                if self.checkBox_4.isChecked():
                    ############################################################################################
                    if self.comboBox != "Virtualization":
                        satir = tarih_hucre.row + 1
                        sutun = tarih_hucre.column
                        value = worksheet.cell(row=satir, column=sutun).value
                        if value is None or len(value) == 1:
                            worksheet.cell(row=satir, column=sutun, value=(text + ' - ' + self.zaman + ' - ' + ' (' + self.isim + ')' + ' - ' + self.comboBox + ' - ' + 'Company: ' + self.combobox2 + ' - ' + 'RESERVE' + ' - '))
                            workbook.save('Planlı Çalışma-Freeze Takvimi.xlsx')
                            QMessageBox.information(self, "Mission status", "The task has been created.")
                        else:
                            QMessageBox.warning(self, "Mission status", "There is work on this date.")

                    elif self.comboBox == "Virtualization":
                        satir = tarih_hucre.row + 2
                        sutun = tarih_hucre.column
                        value = worksheet.cell(row=satir, column=sutun).value
                        if value is None or len(value) == 1:
                            worksheet.cell(row=satir, column=sutun, value=(text + ' - ' + self.zaman + ' - ' + ' (' + self.isim + ')' + ' - ' + 'Company: ' + self.combobox2 + ' - ' + 'RESERVE' + ' - '))
                            workbook.save('Planlı Çalışma-Freeze Takvimi.xlsx')
                            QMessageBox.information(self, "Mission status", "The task has been created.")
                        else:
                            QMessageBox.warning(self, "Mission status", "There is work on this date.")
                    #############################################################################################################
                else:
                    QMessageBox.information(self, "Checkbox Information", "Please tick all checkboxes.")
            #####################################################################################################################


    def upload(self):

        QMessageBox.information(self, "Sharepoint Information","Do not close the browser until the file has been upload.")
        driver = webdriver.Chrome()
        driver.get("https://nortel365.sharepoint.com/sites/CORENETWORKORTAKALANI/Shared Documents/General")
        email = self.lineEdit_2.text()
        şifre = self.lineEdit_3.text()
        ####################################################################################################
        time.sleep(2)
        email_field = driver.find_element(By.CSS_SELECTOR, "input[name='loginfmt']")
        email_field.send_keys(email)
        email_field.send_keys(Keys.RETURN)
        ######################################################################################################
        time.sleep(2)
        password_field = driver.find_element(By.CSS_SELECTOR, "input[name='passwd']")
        password_field.send_keys(şifre)
        password_field.send_keys(Keys.RETURN)
        #####################################################################################################
        time.sleep(2)
        text_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, "//*[text()='Approve a request on my Microsoft Authenticator app']"))
        )
        text_element.click()
        #####################################################################################################
        if EC.presence_of_element_located((By.ID, "idBtn_Back")):
            print("5")
            back_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "idBtn_Back"))
            )
            back_button.click()
        #####################################################################################################
        time.sleep(2)
        text_element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//*[text()='Upload']"))
        )
        text_element.click()
        #####################################################################################################
        time.sleep(2)
        text_element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//*[text()='Files']"))
        )
        text_element.click()

        time.sleep(2)
        #####################################################################################################
        dizin = r"C:\TASK PLANNER"
        dosya_adı = "Planlı Çalışma-Freeze Takvimi.xlsx"
        yol = os.path.join(dizin, dosya_adı)

        keyboard.write(yol)

        keyboard.press_and_release('enter')

 
        time.sleep(2)

        # Yüklemeyi onaylamak için Enter tuşuna basın
        keyboard.press_and_release('enter')

        text_element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//*[text()='Replace']"))
        )
        text_element.click()
        time.sleep(5)
        #####################################################################################################
    def clear(self):
        self.textEdit.clear()



if __name__ == "__main__":
    app = QApplication(sys.argv)
    mainWindow = MainWindow()  
    mainWindow.show()
    sys.exit(app.exec_())

